import streamlit as st
import pandas as pd
import io
import os
import yaml
import json
import smtplib
import string
import secrets as secrets_module
import requests
import base64
import bcrypt
from datetime import datetime, date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit_authenticator as stauth
from yaml.loader import SafeLoader

LOGS_FILE = "logs.json"

# ── Page config ───────────────────────────────────────────────
st.set_page_config(page_title="MusicNet Excel Merger", page_icon="🎵", layout="wide")

# ── CSS ───────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Login card ── */
div[data-testid="stForm"] {
    max-width: 420px;
    margin: 2rem auto 0 auto;
    background: white;
    padding: 2rem 2.5rem 1.5rem 2.5rem;
    border-radius: 16px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.12);
}
div[data-testid="stForm"] h2 { text-align: center; }

/* ── Shrink login inputs ── */
div[data-testid="stTextInput"] input {
    font-size: 0.88rem !important;
    padding: 0.35rem 0.6rem !important;
    height: 36px !important;
}
div[data-testid="stTextInput"] label {
    font-size: 0.85rem !important;
}

/* ── Login button ── */
div[data-testid="stForm"] button[kind="primaryFormSubmit"] {
    width: 100%;
    background: linear-gradient(135deg, #1F3864, #2E75B6) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    margin-top: 0.5rem;
}

/* ── Main app ── */
.main { background-color: #f8f9fb; }
.block-container { padding-top: 1.5rem; }
.title-bar {
    background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
    padding: 1.2rem 2rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
    color: white;
}
.title-bar h1 { color: white; margin: 0; font-size: 1.8rem; }
.title-bar p  { color: #cce0f5; margin: 0.2rem 0 0 0; font-size: 0.95rem; }
.metric-card {
    background: white;
    border-radius: 10px;
    padding: 1.2rem;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    border-left: 4px solid #2E75B6;
}
.metric-card .num { font-size: 2.2rem; font-weight: 700; color: #1F3864; }
.metric-card .lbl { font-size: 0.85rem; color: #666; margin-top: 0.2rem; }
.stButton > button {
    background: linear-gradient(135deg, #1F3864, #2E75B6);
    color: white; border: none; border-radius: 8px;
    padding: 0.6rem 2rem; font-size: 1rem;
    font-weight: 600; width: 100%;
}
.stButton > button:hover { opacity: 0.9; }
/* Compact secondary/pagination buttons */
.stButton > button[kind="secondary"] {
    background: #f0f4fa !important;
    color: #1F3864 !important;
    border: 1px solid #c8d8ed !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    padding: 0.3rem 0.5rem !important;
}
.stButton > button[kind="secondary"]:hover {
    background: #dce8f5 !important;
    border-color: #2E75B6 !important;
}
.success-box {
    background: #e8f5e9; border-left: 4px solid #43a047;
    padding: 1rem 1.2rem; border-radius: 8px;
    color: #2e7d32; font-weight: 600;
}
.section-header {
    font-size: 1.1rem; font-weight: 700; color: #1F3864;
    border-bottom: 2px solid #2E75B6;
    padding-bottom: 0.4rem; margin-bottom: 1rem;
}
.login-title {
    text-align: center;
    color: #1F3864;
    font-size: 1.6rem;
    font-weight: 700;
    margin-bottom: 0.2rem;
}
.login-sub {
    text-align: center;
    color: #888;
    font-size: 0.9rem;
    margin-bottom: 1.5rem;
}
</style>
""", unsafe_allow_html=True)

# ── Load config ───────────────────────────────────────────────
with open("config.yaml") as f:
    config = yaml.load(f, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config["credentials"],
    config["cookie"]["name"],
    config["cookie"]["key"],
    config["cookie"]["expiry_days"],
)

# ── Helper: read logs ────────────────────────────────────────
def read_logs():
    # 1. Try GitHub first (most persistent — survives redeploys)
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO", "gil-hue/musicnet-merger")
        if token:
            url = f"https://api.github.com/repos/{repo}/contents/logs.json"
            r   = requests.get(url, headers={"Authorization": f"token {token}"})
            if r.status_code == 200:
                return json.loads(base64.b64decode(r.json()["content"]).decode())
    except Exception:
        pass
    # 2. Try local file (persists between browser sessions while app is running)
    try:
        if os.path.exists(LOGS_FILE):
            with open(LOGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list) and data:
                    return data
    except Exception:
        pass
    # 3. Fallback: in-memory session state (current session only)
    return st.session_state.get("session_logs", [])

# ── Helper: write log entry ──────────────────────────────────
def write_log(action, details, user=None):
    entry = {
        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "user":      user or st.session_state.get("username", "—"),
        "action":    action,
        "details":   details
    }
    # Layer 1: always save to session state
    if "session_logs" not in st.session_state:
        st.session_state["session_logs"] = []
    st.session_state["session_logs"].insert(0, entry)

    # Layer 2: write to local file (persists between sessions while app is awake)
    try:
        logs = []
        if os.path.exists(LOGS_FILE):
            with open(LOGS_FILE, "r", encoding="utf-8") as f:
                logs = json.load(f)
        if not isinstance(logs, list):
            logs = []
        logs.insert(0, entry)
        logs = logs[:500]
        with open(LOGS_FILE, "w", encoding="utf-8") as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    # Layer 3: also persist to GitHub (survives redeploys — requires GITHUB_TOKEN)
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO", "gil-hue/musicnet-merger")
        if not token:
            return
        url     = f"https://api.github.com/repos/{repo}/contents/logs.json"
        headers = {"Authorization": f"token {token}"}
        r       = requests.get(url, headers=headers)
        if r.status_code != 200:
            return
        sha  = r.json()["sha"]
        logs = json.loads(base64.b64decode(r.json()["content"]).decode())
        logs.insert(0, entry)
        logs = logs[:500]
        content = base64.b64encode(json.dumps(logs, ensure_ascii=False, indent=2).encode()).decode()
        requests.put(url, headers=headers, json={"message": f"Log: {action}", "content": content, "sha": sha})
    except Exception:
        pass

# ── Helper: update config.yaml on GitHub ─────────────────────
def update_config_github(new_config):
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO", "gil-hue/musicnet-merger")
        if not token:
            return False
        url     = f"https://api.github.com/repos/{repo}/contents/config.yaml"
        headers = {"Authorization": f"token {token}"}
        r       = requests.get(url, headers=headers)
        sha     = r.json().get("sha", "")
        content = base64.b64encode(yaml.dump(new_config, allow_unicode=True).encode()).decode()
        requests.put(url, headers=headers, json={
            "message": "Reset user password",
            "content": content,
            "sha": sha
        })
        return True
    except Exception:
        return False

# ── Helper: send email ────────────────────────────────────────
def send_reset_email(to_email, new_password, username):
    try:
        smtp_host = st.secrets.get("SMTP_HOST", "")
        smtp_port = int(st.secrets.get("SMTP_PORT", 587))
        smtp_user = st.secrets.get("SMTP_USER", "")
        smtp_pass = st.secrets.get("SMTP_PASS", "")
        if not all([smtp_host, smtp_user, smtp_pass]):
            return False
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "🎵 MusicNet — איפוס סיסמה"
        msg["From"]    = smtp_user
        msg["To"]      = to_email
        body = f"""
        <div dir="rtl" style="font-family:Arial;max-width:460px;margin:auto;padding:2rem;
             background:#f8f9fb;border-radius:12px;">
          <h2 style="color:#1F3864;">🎵 MusicNet Excel Merger</h2>
          <p>שלום <strong>{username}</strong>,</p>
          <p>הסיסמה שלך אופסה. הסיסמה הזמנית שלך היא:</p>
          <div style="background:#1F3864;color:white;font-size:1.4rem;font-weight:bold;
               letter-spacing:3px;text-align:center;padding:1rem;border-radius:8px;
               margin:1rem 0;">{new_password}</div>
          <p style="color:#888;font-size:0.85rem;">אנא שנה את הסיסמה לאחר הכניסה.</p>
        </div>
        """
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_email, msg.as_string())
        return True
    except Exception:
        return False

# ── Forgot password UI ────────────────────────────────────────
def show_forgot_password():
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("🔑 שכחת סיסמה?"):
        fp_user = st.text_input("שם משתמש", key="fp_user", placeholder="הכנס שם משתמש")
        if st.button("שלח סיסמה חדשה", key="fp_btn"):
            users = config["credentials"]["usernames"]
            if fp_user in users:
                user_email = users[fp_user].get("email", "")
                alphabet   = string.ascii_letters + string.digits
                new_pass   = ''.join(secrets_module.choice(alphabet) for _ in range(10))
                hashed     = bcrypt.hashpw(new_pass.encode(), bcrypt.gensalt()).decode()
                config["credentials"]["usernames"][fp_user]["password"] = hashed
                github_ok  = update_config_github(config)
                email_ok   = send_reset_email(user_email, new_pass, fp_user)
                if email_ok:
                    st.success(f"✅ סיסמה חדשה נשלחה לאימייל {user_email}")
                elif github_ok:
                    st.warning("⚠️ הסיסמה עודכנה אך שליחת המייל נכשלה — פנה לאדמין")
                else:
                    st.info("ℹ️ שירות המייל אינו מוגדר — פנה לאדמין לאיפוס סיסמה")
            else:
                st.error("שם משתמש לא נמצא במערכת")

# ── Login page ────────────────────────────────────────────────
auth_status = st.session_state.get("authentication_status")

if auth_status is not True:
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown('<div class="login-title">🎵 MusicNet Excel Merger</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-sub">כניסה למערכת</div>', unsafe_allow_html=True)
        authenticator.login(location="main")
        if auth_status is False:
            st.error("❌ שם משתמש או סיסמה שגויים")
        show_forgot_password()
    st.stop()

# ── Logged in bar ─────────────────────────────────────────────
col_u, col_lo = st.columns([9, 1])
with col_lo:
    authenticator.logout("התנתק")

# ── Header ────────────────────────────────────────────────────
current_user = st.session_state.get("username", "")
is_admin     = config["credentials"]["usernames"].get(current_user, {}).get("role") == "admin"

st.markdown("""
<div class="title-bar">
    <h1>🎵 MusicNet Excel Merger</h1>
    <p>העלה קבצי Excel, בחר עמודות, וקבל קובץ מאוחד להורדה</p>
</div>
""", unsafe_allow_html=True)

# ── Tabs (admin sees extra tab) ───────────────────────────────
if is_admin:
    tab_main, tab_admin, tab_log = st.tabs(["🎵 מיזוג קבצים", "👥 ניהול משתמשים", "📋 לוג פעולות"])
else:
    tab_main = st.tabs(["🎵 מיזוג קבצים"])[0]
    tab_admin = None

# ══════════════════════════════════════════════════════════════
#  ADMIN TAB
# ══════════════════════════════════════════════════════════════
if is_admin and tab_log:
    with tab_log:
        st.markdown('<div class="section-header">📋 לוג פעולות</div>', unsafe_allow_html=True)
        col_r, col_f = st.columns([1, 3])
        with col_r:
            if st.button("🔄 רענן לוג", key="refresh_log"):
                st.rerun()
        with col_f:
            filter_action = st.selectbox("סנן לפי פעולה", ["הכל", "העלאת קבצים", "יצירת קובץ ממוזג", "הורדת קובץ"], key="log_filter")

        logs = read_logs()
        if logs:
            if filter_action != "הכל":
                logs = [l for l in logs if l.get("action") == filter_action]
            if logs:
                action_icons = {
                    "העלאת קבצים":      "📂",
                    "יצירת קובץ ממוזג": "⚙️",
                    "הורדת קובץ":       "⬇️",
                }
                df_log = pd.DataFrame([{
                    "🕐 תאריך ושעה": l.get("timestamp",""),
                    "👤 משתמש":       l.get("user",""),
                    "פעולה":          action_icons.get(l.get("action",""), "•") + " " + l.get("action",""),
                    "📝 פרטים":       l.get("details","")
                } for l in logs])
                # ── Pagination ──────────────────────────────────────
                ITEMS_PER_PAGE = 20
                total_pages    = max(1, (len(df_log) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
                if st.session_state.get("_last_log_filter") != filter_action:
                    st.session_state["log_page"]         = 0
                    st.session_state["_last_log_filter"] = filter_action
                if "log_page" not in st.session_state:
                    st.session_state["log_page"] = 0
                page = min(st.session_state["log_page"], total_pages - 1)
                st.session_state["log_page"] = page
                start   = page * ITEMS_PER_PAGE
                df_page = df_log.iloc[start : start + ITEMS_PER_PAGE]

                st.dataframe(df_page, use_container_width=True, hide_index=True)
                st.caption(f"סה\"כ: {len(df_log)} פעולות | עמוד {page + 1} מתוך {total_pages}")

                if total_pages > 1:
                    half    = 2
                    p_start = max(0, page - half)
                    p_end   = min(total_pages, p_start + 5)
                    if p_end - p_start < 5:
                        p_start = max(0, p_end - 5)
                    page_range = list(range(p_start, p_end))
                    pcols = st.columns([0.5] + [0.3] * len(page_range) + [0.5], gap="small")
                    with pcols[0]:
                        if st.button("‹", key="pg_prev", disabled=(page == 0), type="secondary"):
                            st.session_state["log_page"] = page - 1
                            st.rerun()
                    for i, pg in enumerate(page_range):
                        with pcols[i + 1]:
                            b_type = "primary" if pg == page else "secondary"
                            if st.button(str(pg + 1), key=f"pg_{pg}", type=b_type):
                                st.session_state["log_page"] = pg
                                st.rerun()
                    with pcols[-1]:
                        if st.button("›", key="pg_next", disabled=(page == total_pages - 1), type="secondary"):
                            st.session_state["log_page"] = page + 1
                            st.rerun()

                st.markdown("---")
                buf = io.BytesIO()
                df_log.to_excel(buf, index=False)
                buf.seek(0)
                st.download_button("⬇️ ייצא לוג ל-Excel", data=buf.read(),
                                   file_name=f"MusicNet_Log_{date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("אין פעולות מסוג זה בלוג")
        else:
            st.info("הלוג ריק — פעולות יופיעו כאן לאחר שימוש באפליקציה")
            st.caption("💡 הלוג מצריך הגדרת GITHUB_TOKEN ב-Secrets")

if is_admin and tab_admin:
    with tab_admin:
        st.markdown('<div class="section-header">👥 ניהול משתמשים</div>', unsafe_allow_html=True)
        users = config["credentials"]["usernames"]

        # ── Users table ───────────────────────────────────────
        rows = [{"📧 אימייל / שם משתמש": u,
                 "👤 שם": d.get("name",""),
                 "🔑 תפקיד": "👑 אדמין" if d.get("role")=="admin" else "👤 משתמש"}
                for u, d in users.items()]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        st.markdown(f"**סה\"כ: {len(users)} משתמשים**")
        st.markdown("---")

        # ── Add user ──────────────────────────────────────────
        st.markdown("#### ➕ הוספת משתמש חדש")
        c1, c2 = st.columns(2)
        with c1:
            new_email = st.text_input("אימייל", key="nu_email", placeholder="user@example.com")
            new_name  = st.text_input("שם מלא", key="nu_name",  placeholder="ישראל ישראלי")
        with c2:
            new_pass  = st.text_input("סיסמה", key="nu_pass", type="password", placeholder="לפחות 6 תווים")
            new_role  = st.selectbox("תפקיד", ["user", "admin"], key="nu_role",
                                     format_func=lambda x: "👑 אדמין" if x=="admin" else "👤 משתמש")
        if st.button("➕ הוסף משתמש", key="btn_add"):
            if not new_email or not new_pass or not new_name:
                st.error("יש למלא את כל השדות")
            elif new_email in users:
                st.error("המשתמש כבר קיים במערכת")
            elif len(new_pass) < 6:
                st.error("הסיסמה חייבת להכיל לפחות 6 תווים")
            else:
                hashed = bcrypt.hashpw(new_pass.encode(), bcrypt.gensalt()).decode()
                config["credentials"]["usernames"][new_email] = {
                    "email": new_email, "name": new_name,
                    "password": hashed, "role": new_role
                }
                if update_config_github(config):
                    st.success(f"✅ המשתמש {new_email} נוסף בהצלחה!")
                    st.rerun()
                else:
                    st.warning("⚠️ המשתמש נוסף אך לא נשמר — הגדר GITHUB_TOKEN ב-Secrets")

        st.markdown("---")

        # ── Reset password ────────────────────────────────────
        st.markdown("#### 🔄 איפוס סיסמה למשתמש")
        other_users = [u for u in users if u != current_user]
        if other_users:
            reset_user = st.selectbox("בחר משתמש", other_users, key="reset_user")
            c1, c2 = st.columns(2)
            with c1:
                manual_pass = st.text_input("סיסמה חדשה (השאר ריק לסיסמה אוטומטית)",
                                            key="reset_pass", type="password")
            with c2:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("🔄 אפס סיסמה", key="btn_reset"):
                    alphabet = string.ascii_letters + string.digits
                    new_p    = manual_pass if manual_pass else ''.join(secrets_module.choice(alphabet) for _ in range(10))
                    hashed   = bcrypt.hashpw(new_p.encode(), bcrypt.gensalt()).decode()
                    config["credentials"]["usernames"][reset_user]["password"] = hashed
                    github_ok = update_config_github(config)
                    email_ok  = send_reset_email(users[reset_user].get("email",""), new_p, reset_user)
                    if email_ok:
                        st.success(f"✅ סיסמה חדשה נשלחה לאימייל של {reset_user}")
                    else:
                        st.info(f"✅ הסיסמה החדשה: **`{new_p}`** — שלח למשתמש ידנית")
        else:
            st.info("אין משתמשים נוספים במערכת")

        st.markdown("---")

        # ── Delete user ───────────────────────────────────────
        st.markdown("#### 🗑️ מחיקת משתמש")
        if other_users:
            del_user = st.selectbox("בחר משתמש למחיקה", other_users, key="del_user")
            if st.button("🗑️ מחק משתמש", key="btn_del", type="primary"):
                if st.session_state.get("confirm_del") != del_user:
                    st.session_state["confirm_del"] = del_user
                    st.warning(f"⚠️ לחץ שוב לאישור מחיקת **{del_user}**")
                else:
                    del config["credentials"]["usernames"][del_user]
                    st.session_state.pop("confirm_del", None)
                    if update_config_github(config):
                        st.success(f"✅ המשתמש {del_user} נמחק")
                        st.rerun()
                    else:
                        st.warning("⚠️ נמחק מקומית אך לא נשמר — הגדר GITHUB_TOKEN")
        else:
            st.info("אין משתמשים נוספים למחיקה")

# ── Helper: build Excel ───────────────────────────────────────
def build_excel(summary_data, merged_df, col_headers):
    wb = Workbook()
    header_fill = PatternFill("solid", start_color="1F3864")
    sub_fill    = PatternFill("solid", start_color="2E75B6")
    total_fill  = PatternFill("solid", start_color="D6E4F0")
    white_fill  = PatternFill("solid", start_color="FFFFFF")
    alt_fill    = PatternFill("solid", start_color="EBF2FA")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    col_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_font   = Font(name="Arial", size=11)
    total_font  = Font(name="Arial", bold=True, size=11, color="1F3864")
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    ws_s = wb.active
    ws_s.title = "סיכום"
    ws_s.merge_cells("A1:B1")
    ws_s["A1"] = "סיכום קבצי Excel שעובדו"
    ws_s["A1"].font = header_font; ws_s["A1"].fill = header_fill; ws_s["A1"].alignment = center
    ws_s.row_dimensions[1].height = 32
    ws_s.merge_cells("A2:B2")
    ws_s["A2"] = "תאריך עיבוד: " + date.today().strftime("%d/%m/%Y")
    ws_s["A2"].font = Font(name="Arial", italic=True, size=10, color="595959"); ws_s["A2"].alignment = center
    ws_s.row_dimensions[2].height = 18
    for col, txt in [(1, "שם הקובץ"), (2, "מספר רשומות")]:
        c = ws_s.cell(3, col, txt)
        c.font = col_font; c.fill = sub_fill; c.alignment = center; c.border = border
    ws_s.row_dimensions[3].height = 24
    for i, (fname, cnt) in enumerate(summary_data):
        r = i + 4
        fill = white_fill if i % 2 == 0 else alt_fill
        c1 = ws_s.cell(r, 1, fname); c2 = ws_s.cell(r, 2, cnt)
        for c in [c1, c2]: c.font = data_font; c.fill = fill; c.border = border
        c1.alignment = left; c2.alignment = center
        ws_s.row_dimensions[r].height = 22
    tr = len(summary_data) + 4
    c1 = ws_s.cell(tr, 1, 'סה"כ רשומות'); c2 = ws_s.cell(tr, 2, "=SUM(B4:B" + str(tr-1) + ")")
    for c in [c1, c2]: c.font = total_font; c.fill = total_fill; c.border = border
    c1.alignment = left; c2.alignment = center
    ws_s.row_dimensions[tr].height = 24
    ws_s.column_dimensions["A"].width = 35; ws_s.column_dimensions["B"].width = 20

    ws_d = wb.create_sheet("נתונים מאוחדים")
    for ci, col in enumerate(merged_df.columns, 1):
        c = ws_d.cell(1, ci, col_headers.get(col, col))
        c.font = col_font; c.fill = sub_fill; c.alignment = center; c.border = border
        ws_d.column_dimensions[get_column_letter(ci)].width = 32
    ws_d.row_dimensions[1].height = 26
    for ri, row in enumerate(merged_df.itertuples(index=False), 2):
        fill = white_fill if ri % 2 == 0 else alt_fill
        for ci, val in enumerate(row, 1):
            c = ws_d.cell(ri, ci, str(val) if pd.notna(val) else "")
            c.font = data_font; c.fill = fill; c.alignment = left; c.border = border
        ws_d.row_dimensions[ri].height = 18
    ws_d.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════════════════════
#  MAIN UI (tab)
# ══════════════════════════════════════════════════════════════
with tab_main:

    # ── Step 1: Upload files ──────────────────────────────────
    st.markdown('<div class="section-header">📂 שלב 1 — העלאת קבצים</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("גרור קבצי Excel לכאן", type=["xlsx","xls"],
                                 accept_multiple_files=True, label_visibility="collapsed")
    if uploaded:
        file_data   = {}
        all_columns = set()
        upload_time = datetime.now().strftime("%d/%m/%Y %H:%M")
        for f in uploaded:
            try:
                df = pd.read_excel(f)
                file_data[f.name] = {"df": df, "upload_time": upload_time}
                all_columns.update(df.columns.tolist())
            except Exception as e:
                st.error(f"שגיאה בקריאת {f.name}: {e}")
        st.session_state["file_data"]   = file_data
        st.session_state["all_columns"] = sorted(all_columns)
        log_key = f"logged_upload_{sorted(file_data.keys())}"
        if file_data and not st.session_state.get(log_key):
            total_rows = sum(len(v["df"]) for v in file_data.values())
            write_log("העלאת קבצים",
                      f"{len(file_data)} קבצים | {total_rows:,} רשומות: {', '.join(file_data.keys())}")
            st.session_state[log_key] = True
            st.session_state.pop("excel_result", None)

    # Load from session_state (persists across tab switches)
    file_data   = st.session_state.get("file_data",   {})
    all_columns = st.session_state.get("all_columns", [])

    if not file_data:
        st.info("⬆️ העלה לפחות קובץ Excel אחד כדי להמשיך.")
    else:
        total_recs = sum(len(v["df"]) for v in file_data.values())
        c1, c2, c3 = st.columns(3)
        with c1: st.markdown(f'<div class="metric-card"><div class="num">{len(file_data)}</div><div class="lbl">קבצים הועלו</div></div>', unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="metric-card"><div class="num">{total_recs:,}</div><div class="lbl">רשומות סה"כ</div></div>', unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="metric-card"><div class="num">{len(all_columns)}</div><div class="lbl">עמודות נמצאו</div></div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        with st.expander("📋 פירוט רשומות לפי קובץ", expanded=True):
            rows = [{"📄 שם קובץ": name, "📅 תאריך העלאה": v["upload_time"],
                     "📊 רשומות": len(v["df"]), "📑 עמודות": len(v["df"].columns)}
                    for name, v in file_data.items()]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
                         column_config={"📊 רשומות": st.column_config.NumberColumn(format="%d")})

        # ── Step 2: Column selection ──────────────────────────
        st.markdown('<div class="section-header">🔧 שלב 2 — בחירת עמודות לשמור</div>', unsafe_allow_html=True)
        default_cols  = [c for c in ['album_name','artist_name','track_name','track_uri','label'] if c in all_columns]
        selected_cols = st.multiselect("בחר את העמודות שברצונך לשמור:", options=all_columns, default=default_cols)

        if selected_cols:
            # ── Step 3: Process ───────────────────────────────
            st.markdown('<div class="section-header">⚙️ שלב 3 — עיבוד ויצוא</div>', unsafe_allow_html=True)
            col_headers = {'album_name':'שם אלבום','artist_name':'שם אמן',
                           'track_name':'שם שיר','track_uri':'Track URI','label':'לייבל'}

            if st.button("🚀 עבד וצור קובץ מאוחד", use_container_width=True):
                with st.spinner("מעבד קבצים..."):
                    summary_data = []
                    dfs = []
                    for fname, v in file_data.items():
                        df       = v["df"]
                        existing = [c for c in selected_cols if c in df.columns]
                        missing  = [c for c in selected_cols if c not in df.columns]
                        dfs.append(df[existing])
                        summary_data.append((fname, len(df)))
                        if missing: st.warning(f"⚠️ {fname}: עמודות חסרות — {missing}")
                    merged      = pd.concat(dfs, ignore_index=True)
                    today       = date.today().strftime("%Y-%m-%d")
                    out_name    = f"MusicNet_Merged_{today}.xlsx"
                    excel_bytes = build_excel(summary_data, merged, col_headers)
                st.session_state["excel_result"] = {
                    "bytes":   excel_bytes,
                    "name":    out_name,
                    "count":   len(merged),
                    "cols":    len(merged.columns),
                    "files":   len(file_data),
                    "preview": merged.head(20).to_dict("records"),
                }
                write_log("יצירת קובץ ממוזג",
                          f"{out_name} | {len(merged):,} רשומות מ-{len(file_data)} קבצים")

            # ── Results panel (persists across tab switches) ──
            if "excel_result" in st.session_state:
                er = st.session_state["excel_result"]
                st.markdown(
                    f'<div class="success-box">✅ הקובץ המאוחד נוצר בהצלחה!<br>'
                    f'סה"כ רשומות: <strong>{er["count"]:,}</strong> | '
                    f'עמודות: <strong>{er["cols"]}</strong></div>',
                    unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                if st.download_button(
                    label="⬇️ הורד את הקובץ המאוחד",
                    data=er["bytes"],
                    file_name=er["name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                ):
                    write_log("הורדת קובץ", f"{er['name']} | {er['count']:,} רשומות")
                with st.expander("📊 תצוגה מקדימה — 20 שורות ראשונות"):
                    st.dataframe(pd.DataFrame(er["preview"]),
                                 use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ יש לבחור לפחות עמודה אחת.")
